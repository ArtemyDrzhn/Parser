import openpyxl


def export(comment, name):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)

    sheet = wb['Первый лист']
    sheet.append([name])
    for i in comment:
        sheet.append(i)
    name_xlsx = name.replace('https://vk.com/', '') + '.xlsx'
    wb.save('data//' + name_xlsx)


def read_xlsx(path):
    """
    Чтение из экселя
    :param path: путь до файла
    :return: дикт (кей = ком, вал = лейбл)
    """
    sheet = openpyxl.load_workbook(path).active
    counter = 1
    lst_1 = []
    while sheet.cell(counter, 1).value is not None:
        lst_1.append(sheet.cell(counter, 1).value)
        counter += 1

    counter = 1
    lst_2 = []
    while sheet.cell(counter, 2).value is not None:
        lst_2.append(sheet.cell(counter, 2).value)
        counter += 1
    d = {}
    if len(lst_1) == len(lst_2):
        # d = dict(zip(lst_1, lst_2))
        for i in lst_1:
            d[str(i)] = lst_2
    else:
        print(f"Кол-во лейблов не совпадает с кол-ом комментов в файле {path}")

    # Список кортежей: 1 - коммент, 2 - лейбл
    unif = list(zip(lst_1, lst_2))
    return unif


# TODO Нужно разбить на три класса - плохой, средний, положительный
# 0-это плохо, 1-нейтрально, 2-положительно
def corpus_write(unif):
    """
    Создание корпуса данных
    :param unif: список кортежей
    :return:
    """
    bad = open('corpus_root//bad.txt', 'w', encoding='utf-8')
    neutral = open('corpus_root//neutral.txt', 'w', encoding='utf-8')
    good = open('corpus_root//good.txt', 'w', encoding='utf-8')
    for i in unif:
        if i[1] == 0:
            bad.write(i[0])
            bad.write('\n')
        elif i[1] == 1:
            neutral.write(i[0])
            neutral.write('\n')
        elif i[1] == 2:
            good.write(i[0])
            good.write('\n')
